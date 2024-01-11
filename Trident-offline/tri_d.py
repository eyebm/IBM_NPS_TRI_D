from datetime import datetime
import pandas as pd
import numpy as np
import l1, l2, gui, migration
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import os, sys
from cpu_disk_burn import cpu_disk_burn
from tkinter import DISABLED
import tkinter as tk


def main():
    version = "1.0.4"
    warnings.simplefilter(action='ignore', category=FutureWarning)
    gui.create_window(version)

def run_trid(inpath, outpath, percent, log, window):


    log.config(state=tk.NORMAL)
    log.delete('1.0', tk.END)            

    log.insert(tk.END, 'Starting Tri-D Processing...\n')
    window.update_idletasks()
    inpath += "/"
    outpath += "/" 
    if percent == "":
        percent = .90
    else:
        percent = float(percent.strip('%'))/100
    level_1(inpath, outpath, log)
    window.update_idletasks()
    user_level(inpath, outpath, percent, log)
    window.update_idletasks()
    log.see("end")
    burn(inpath, outpath, log)
    window.update_idletasks()
    log.see("end")
    migration_analysis(inpath,outpath,log)
    window.update_idletasks()
    log.see("end")
    log.insert(tk.END, "Tri-D Complete!")
    window.update_idletasks()
    log.see("end")
    log.config(state=DISABLED)

def level_1(inpath, outpath, log):
    mode = 'INT'
    ext = '.csv'

    if not os.path.exists(inpath + "int-query-day.csv") and not os.path.exists(inpath + "int-query-yr.csv") and not os.path.exists(inpath + "int-query-dow.csv"):
        if not os.path.exists(inpath + "int-query-day.txt") and not os.path.exists(inpath + "int-query-yr.txt") and not os.path.exists(inpath + "int-query-dow.txt"):
            if not os.path.exists(inpath + "qh-query-day.txt") and not os.path.exists(inpath + "qh-query-yr.txt") and not os.path.exists(inpath + "qh-query-dow.txt"):
                if not os.path.exists(inpath + "qh-query-day.csv") and not os.path.exists(inpath + "qh-query-yr.csv") and not os.path.exists(inpath + "qh-query-dow.csv"):
                    log.insert(tk.END, "Level 1: Input files not found, skipping Level 1 Analysis\n")
                    return
                else:
                    mode = 'QH'
            else:
                mode = 'QH'
                ext = '.txt'
        else:
            ext = '.txt'
    

    wb = Workbook()
    ws = wb.active
    if mode == 'INT':
        
        if not os.path.exists(inpath + "int-query-day" + ext):
            log.insert(tk.END, "Level 1: 'int-query-day' not found. Skipping Month by Hour.\n")
        elif not os.path.exists(inpath + "int-query-yr" + ext):
            log.insert(tk.END, "Level 1: 'int-query-yr' not found. Skipping Month by Hour.\n")
        else:
            ws.title = "Month by Hour"
            data = pd.read_csv(inpath + "int-query-day" + ext, sep="|")
            if not len(data.index):
                log.insert(tk.END, "Level 1: No data in 'int-query-day'. Skipping Month by Hour.\n")
            else:
                data.columns = map(lambda x: str.upper(x), data.columns) 
                yr_data = pd.read_csv(inpath + "int-query-yr" + ext, sep = "|")
                yr_data.columns = map(lambda x: str.upper(x), yr_data.columns)
                if not len(data.index):
                    log.insert(tk.END, "Level 1: No data in 'int-query-yr'. Skipping Month by Hour.\n")
                else: 
                    result = l1.process_yr(data, yr_data)
                    ws_append(ws, result)

        if not os.path.exists(inpath + "int-query-day" + ext):
            log.insert(tk.END, "Level 1: 'int-query-day' not found. Skipping Hour by Hour Avg.\n")
        else: 
            ws = wb.create_sheet("Hour by Hour Avg")
            data = pd.read_csv(inpath + "int-query-day" + ext, sep = '|')
            if not len(data.index):
                log.insert(tk.END, "Level 1: No data in 'int-query-day'. Skipping Hour by Hour Avg.\n")
            else:
                data.columns = map(lambda x: str.upper(x), data.columns)
                result = l1.process(data, "HR")
                result = l2.fill_missing(result, range(24))
                result = result.sort_values(by = 'HR')
                ws_append(ws, result)
        
        if not os.path.exists(inpath + "int-query-dow" + ext):
            log.insert(tk.END, "Level 1: 'int-query-day' not found. Skipping DOW by Hour.\n")    
        else:    
            ws = wb.create_sheet("DOW by Hour")
            data = pd.read_csv(inpath + "int-query-dow" + ext, sep = '|')
            data = l1.trim_space(data)
            if not len(data.index):
                log.insert(tk.END, "Level 1: No data in 'int-query-dow'. Skipping DOW by Hour.\n")
            else:
                data.columns = map(lambda x: str.upper(x), data.columns)
                data['WEEKDAY'] = data['DT'].apply(lambda x: pd.to_datetime(x).day_name())
                result = l1.process(data, "WEEKDAY")
                result = l2.fill_missing(result, ['Sunday', 'Monday', 'Tuesday','Wednesday', 'Thursday', 'Friday', 'Saturday'])
                result = sort_weekday(result)
                ws_append(ws, result)
    else:
        if not os.path.exists(inpath + "qh-query-day" + ext):
            log.insert(tk.END, "Level 1: 'qh-query-day' not found. Skipping Month by Hour.\n")
        elif not os.path.exists(inpath + "qh-query-yr" + ext):
            log.insert(tk.END, "Level 1: 'qh-query-yr' not found. Skipping Month by Hour.\n")
        else:
            ws.title = "Month by Hour"
            data = pd.read_csv(inpath + "qh-query-day" + ext, sep="|")
            data = l1.trim_space(data)
            if not len(data.index):
                log.insert(tk.END, "Level 1: No data in 'qh-query-day'. Skipping Month by Hour.\n")
            else:
                data.columns = map(lambda x: str.upper(x), data.columns)
                if 'SNIPS' in data.columns:
                    data = data.drop(columns='SNIPS')
                yr_data = pd.read_csv(inpath + "qh-query-yr" + ext, sep = "|")
                yr_data.columns = map(lambda x: str.upper(x), yr_data.columns)
                yr_data = l1.trim_space(yr_data)
                if not len(yr_data.index):
                    log.insert(tk.END, "Level 1: No data in 'qh-query-yr'. Skipping Month by Hour.\n")
                else:
                    yr_data = yr_data.rename(columns={'YEAR':"YR"})
                    yr_data = yr_data.drop(columns='SNIPS')
                    result = l1.process_yr(data, yr_data)
                    ws_append(ws, result)

        if not os.path.exists(inpath + "qh-query-day" + ext):
            log.insert(tk.END, "Level 1: 'qh-query-day' not found. Skipping Hour by Hour Avg.\n")
        else: 
            ws = wb.create_sheet("Hour by Hour Avg")
            data = pd.read_csv(inpath + "qh-query-day" + ext, sep = '|')
            data = l1.trim_space(data)
            if not len(data.index):
                log.insert(tk.END, "Level 1: No data in 'qh-query-day'. Skipping Hour by Hour Avg.\n")
            else:
                data.columns = map(lambda x: str.upper(x), data.columns)
                if 'SNIPS' in data.columns:
                    data = data.drop(columns='SNIPS')
                result = l1.process(data, "HR")
                ws_append(ws, result)
        # result.to_csv("l1_hr.csv", sep="\t")
        if not os.path.exists(inpath + "qh-query-dow" + ext):
            log.insert(tk.END, "Level 1: 'qh-query-dow' not found. Skipping DOW by Hour.\n")    
        else:    
            ws = wb.create_sheet("DOW by Hour")
            data = pd.read_csv(inpath + "qh-query-dow" + ext, sep = '|')
            data = l1.trim_space(data)
            if not len(data.index):
                log.insert(tk.END, "Level 1: No data in 'qh-query-dow'. Skipping DOW by Hour.\n")
            else:
                data.columns = map(lambda x: str.upper(x), data.columns)
                data = data.drop(columns='SNIPS')
                data = data.rename(columns={'DOW':'WEEKDAY'})
                days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
                data['WEEKDAY'] = data['WEEKDAY'].apply(lambda x: days[x-1])
                result = l1.process(data, "WEEKDAY").drop(columns= 'CNT')
                result = sort_weekday(result)
                ws_append(ws, result)
    wb.save(outpath + 'L1.xlsx')
    log.insert(tk.END, "Level 1: Complete\n")
    


def ws_append(ws, data):
    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

def user_level(inpath, outpath, percent: float, log):
    mode = "INT"
    ext = '.csv'
    if not os.path.exists(inpath + "int-query-by-user.csv"):
        if not os.path.exists(inpath + "int-query-by-user.txt"):
            if not os.path.exists(inpath + "qh-query-by-user.txt"):
                if not os.path.exists(inpath + "qh-query-by-user.csv"):
                    log.insert(tk.END, "User Level: User query data not found. Skipping User Level Analysis.\n")
                    return
                else:
                    mode = "QH"
            else:
                mode = 'QH'
                ext = '.txt'
        else:
            ext = '.txt'
    if mode =="INT":
        data = pd.read_csv(inpath + "int-query-by-user" + ext, sep= "|")
        data = l1.trim_space(data)
    else:
        data = pd.read_csv(inpath + "qh-query-by-user" + ext,sep= "|", dtype={})
        data.columns = [col.strip() for col in list(data.columns)]
        data['CNT'] = pd.to_numeric(data['CNT'])
        data = l1.trim_space(data)
        data = data.rename(columns={'QH_USER':'USERNAME'})
    
    data.columns = map(lambda x: str.upper(x), data.columns)
    data['USERNAME'] = data['USERNAME'].apply(lambda x: str(x))
    if not len(data):
        log.insert(tk.END, "User Level: No data in User query data file. Skipping User Level Analysis.\n")
        return

    frames, unique = l2.top_users(data, percent)

    wb = Workbook()
    ws = wb.active
    ws.title = "TOPS_BY_RTIME"

    ws_append(ws, frames['by_rt'])

    ws = wb.create_sheet("TOPS_BY_QTIME")
    ws_append(ws, frames['by_qt'])

    ws = wb.create_sheet("TOPS_BY_VOLUME")
    ws_append(ws, frames['by_cnt'])

    ws = wb.create_sheet("TOP USERS")
    ws_append(ws, unique)
    months = data['DT'].apply(lambda x: datetime.strptime(x, "%Y-%m-%d")).apply(lambda x: str(x.month) +"-"+ x.strftime("%y")).unique().tolist()
    for user in unique['USERNAME']:
        ws = wb.create_sheet(user)
        result = l2.process_user(data, user, months)
        ws_append(ws, result['dow'])
        ws.append([])
        # ws.insert_rows(ws.max_row+1, 2)
        ws_append(ws, result['hr'])
        ws.append([])
        # ws.insert_rows(ws.max_row+1, 2)
        ws_append(ws, result['yr'])

    wb.save(outpath + "User_Level.xlsx")
    log.insert(tk.END, "User Level: Complete\n")

def burn(inpath, outpath, log):
    if not os.path.exists(inpath+"util-cpu-hour.csv"):
        log.insert(tk.END, "Burn: util-cpu-hour.csv not found. Skipping Burn Analysis.\n")
        return
    data = pd.read_csv(inpath+"util-cpu-hour.csv", sep="|")
    data = l1.trim_space(data)
    if not len(data):
        log.insert(tk.END, "Burn: No data in util-cpu-hour.csv. Skipping Burn Analysis.\n")
        return
    data.columns = map(lambda x: str.upper(x), data.columns)
    result = cpu_disk_burn(data)
    wb = Workbook()
    ws = wb.active
    ws.title = ("BURN_HOUR")
    ws_append(ws, result['by_hr'])

    ws = wb.create_sheet('BURN_DOW')
    ws_append(ws, result['by_dow'])

    ws = wb.create_sheet("BURN MONTH")
    ws_append(ws, result['by_month'])

    wb.save(outpath + "Burn.xlsx")
    log.insert(tk.END, "Burn: Complete\n")

def migration_analysis(inpath, outpath, log):
    if not os.path.exists(inpath+"nz_objects.csv"):
        log.insert(tk.END, "Migration Analysis: nz_objects.csv not found. Skipping Migration Cost Analysis.\n")
        return
    migration.migration_objects(inpath, outpath)
    log.insert(tk.END, "Migration Analysis: Complete.\n")

def sort_weekday(df):
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    df['daynum'] = df['WEEKDAY'].apply(lambda x: days.index(x))
    df = df.sort_values(by='daynum', ignore_index=True)
    df = df.drop(columns='daynum')
    return df

if __name__ == "__main__":
    main()